# 1. packages
import os
import re
import shutil
import numpy as np
import pandas as pd
from tqdm import tqdm 

from openpyxl.styles import Font
from openpyxl import load_workbook

# 2. consonants
random_state = 42
np.random.seed(random_state)
models = ["OLS", "Ridge", "BayesRidge", "Lasso", "Elastic", "Huber", "SVR", "GPR",
          "KNN", "RF", "ET", "GBR", "ABR", "XGB", "2-layer MLP", "3-layer MLP"]
features = ["AcouPros", "m-HuBERT", "Concat"]
symptoms = ['P1', 'P2', 'P3', 'N1', 'N4', 'N6', 'G5', 'G9']

os.makedirs('results', exist_ok=True)
os.makedirs('results/model_comparison/', exist_ok=True)

# 3. functions
def format_sheet_openpyxl(ws, float_cols=None, pct_cols=None):
    """
    Apply formatting to any Excel sheet:
    - Freeze panes at D2
    - Bold header row
    - Apply number formats
    - Add autofilter
    float_cols: list of column letters to format as float (3 decimals)
    pct_cols: list of column letters to format as percentage (1 decimal)
    """
    header_font = Font(name='Calibri', size=11, bold=True)
    body_font   = Font(name='Calibri', size=11)

    ws.freeze_panes = 'D2'

    # Auto column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value)
                if len(val) > max_length:
                    max_length = len(val)
            except:
                pass
        ws.column_dimensions[col_letter].width = max(max_length+2, 10)

    max_row = ws.max_row
    max_col = ws.max_column

    # Header
    for c in range(1, max_col+1):
        ws.cell(row=1, column=c).font = header_font

    # Body
    for r_ in range(2, max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r_, column=c)
            cell.font = body_font
            col_letter = cell.column_letter
            if float_cols and col_letter in float_cols:
                cell.number_format = '0.000'
            if pct_cols and col_letter in pct_cols:
                cell.number_format = '0.0%'

    # Autofilter
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=max_col).coordinate}{max_row}"  



# 4. retrieve information from the excel files
dfs = {}
for symptom in tqdm(symptoms, desc='reading data'):
    
    # model performance 
    records = [pd.read_excel('PCA/results/summary.xlsx', sheet_name=symptom),
                pd.read_excel('MLP_2layer/results/summary.xlsx', sheet_name=symptom),
                pd.read_excel('MLP_3layer/results/summary.xlsx', sheet_name=symptom)]

    df = pd.concat(records)

    # sort dataframe
    df = df.sort_values(
        by=['model', 'feature'], 
        key=lambda x: x.map({v: i for i, v in enumerate(models if x.name=='model' else features)})
    ).reset_index(drop=True)
    
    dfs[symptom] = df

# 5. select the best model per model and feature
# generalization gap: https://arxiv.org/pdf/2209.03032
best_dfs = {}
for symptom in tqdm(symptoms, desc='parameter tuning'):
    df = dfs[symptom]
    df['gen_gap_val'] = (df['val_rmse'] - df['train_rmse']) / df['train_rmse']
    combos = df[['model', 'feature']].drop_duplicates()
    
    df = df[(df['gen_gap_val'] < 0.5) & (df['gen_gap_val'] > 0)]
    
    best_df = df.loc[
        df.groupby(['model', 'feature'])['val_rmse'].idxmin()
    ].reset_index(drop=True)
    
    best_df = combos.merge(best_df, on=['model', 'feature'], how='left')
    best_df = best_df.fillna(np.nan)
    
    best_df = best_df.sort_values(
        by=['model', 'feature'], 
        key=lambda x: x.map({v: i for i, v in enumerate(models if x.name=='model' else features)})
    ).reset_index(drop=True)

    best_dfs[symptom] = best_df

# write into an excel file
with pd.ExcelWriter('results/model_comparison/model_feature_performance.xlsx', engine='openpyxl') as writer:
    for symptom in symptoms:
        best_dfs[symptom].to_excel(writer, sheet_name=symptom, index=False) 

# format
wb = load_workbook('results/model_comparison/model_feature_performance.xlsx')
for sheet in symptoms:
    ws = wb[sheet]

    float_cols = []
    pct_cols = []
    for idx, col_name in enumerate(df.columns):
        col_letter = ws.cell(row=1, column=idx+1).column_letter
        if 'rmse' in col_name or 'r2' in col_name:
            float_cols.append(col_letter)
        elif 'gen_gap' in col_name:
            pct_cols.append(col_letter)
    format_sheet_openpyxl(ws, float_cols=float_cols, pct_cols=pct_cols)

wb.save('results/model_comparison/model_feature_performance.xlsx')

# 5. select the best model per item 
if os.path.exists("results/model_sel/"):
    shutil.rmtree("results/model_sel/")
os.makedirs("results/model_sel/", exist_ok=True) 

best_models = []
for symptom in tqdm(symptoms):
    best_df = best_dfs[symptom].copy()
    best_df['gen_gap_test'] = (best_df['test_rmse'] - best_df['train_rmse']) / best_df['train_rmse']

    candidates = best_df[(best_df['gen_gap_test'] > 0) & (best_df['gen_gap_test'] < 0.5)]
    if candidates.empty:
        candidates = best_df[best_df['test_rmse'].notna()].copy()

    best_model = candidates.loc[candidates['test_rmse'].idxmin()]

    # info to retrieve the best models
    sympt = best_model['symptom']
    model = best_model['model']
    featr = best_model['feature']

    param_str = best_model.get('parameter')
    try:
        param = eval(param_str) if isinstance(param_str, str) and param_str.strip() else {}
    except Exception:
        param = {}

    if model == '3-layer MLP':
        param_str_core = '_'.join(f'{k}-{v}' for k, v in param.items() if k != 'early_epoch')
        src_path = (f"MLP_3layer/results/models/model-MLP3_item-{sympt}_feat-{featr}_param-"
                    f"{{{param_str_core}}}.pth")
        dst_path = f"results/model_sel/{symptom}.pth"
        shutil.copy(src_path, dst_path)

    elif model == '2-layer MLP':
        param_str_core = '_'.join(f'{k}-{v}' for k, v in param.items() if k != 'early_epoch')
        src_path = (f"MLP_2layer/results/models/model-MLP2_item-{sympt}_feat-{featr}_param-"
                    f"{{{param_str_core}}}.pth")
        dst_path = f"results/model_sel/{symptom}.pth"
        shutil.copy(src_path, dst_path)

    else:
        safe_parts = []
        for k, v in param.items():
            v_str = str(v)
            v_safe = re.sub(r'[^\w\-]+', '_', v_str)
            safe_parts.append(f"regressor__{k}-{v_safe}")
        safe_param = "_".join(safe_parts) if safe_parts else "default"
        src_path = f"PCA/results/models/model-{model}_item-{sympt}_feat-{featr}_param-{safe_param}.joblib"
        dst_path = f"results/model_sel/{symptom}.joblib"
        shutil.copy(src_path, dst_path)

    best_models.append(best_model)

best_models_df = pd.DataFrame(best_models).reset_index(drop=True)
best_models_df.to_csv('results/model_comparison/best_model.csv', index=False)

