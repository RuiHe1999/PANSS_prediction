# 1. packages
import os
import glob
import numpy as np
import pandas as pd
from tqdm import tqdm 

from openpyxl.styles import Font
from openpyxl import load_workbook

# 2. functions
def format_sheet_openpyxl(ws, float_cols=None, pct_cols=None):
    """
    Apply formatting to any Excel sheet:
    - Freeze panes at D2
    - Bold header row
    - Apply number formats
    - Add autofilter
    float_cols: list of column letters to format as float (3 decimals)
    pct_cols: list of column letters to format as percentage (1 decimal)
    """
    header_font = Font(name='Calibri', size=11, bold=True)
    body_font   = Font(name='Calibri', size=11)

    ws.freeze_panes = 'D2'

    # Auto column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value)
                if len(val) > max_length:
                    max_length = len(val)
            except:
                pass
        ws.column_dimensions[col_letter].width = max(max_length+2, 10)

    max_row = ws.max_row
    max_col = ws.max_column

    # Header
    for c in range(1, max_col+1):
        ws.cell(row=1, column=c).font = header_font

    # Body
    for r in range(2, max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            col_letter = cell.column_letter
            if float_cols and col_letter in float_cols:
                cell.number_format = '0.000'
            if pct_cols and col_letter in pct_cols:
                cell.number_format = '0.0%'

    # Autofilter
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=max_col).coordinate}{max_row}"  

# 3. consonants
random_state = 42
np.random.seed(random_state)
models = ["OLS", "Ridge", "BayesRidge", "Lasso", "Elastic", "Huber", "SVR",
          "GPR", "KNN", "RF", "ET", "GBR", "ABR", "XGB"]
features = ["AcouPros", "m-HuBERT", "Concat"]
symptoms = ['P1', 'P2', 'P3', 'N1', 'N4', 'N6', 'G5', 'G9']
os.makedirs('results', exist_ok=True)

# 4. retrieve informaiton from the csv files
files = glob.glob("results/model_selection/*.csv")

keep_cols = ['symptom', 'model', 'feature', 'train_rmse', 'val_rmse', 'test_rmse']
drop_cols = ['train_r2', 'val_r2', 'test_r2']

datas = []
for f in files:
    data = pd.read_csv(f).drop(columns=drop_cols)
    
    param_cols = [c for c in data.columns if c not in keep_cols]
    if param_cols:
        data['parameter'] = [{k.replace('regressor__', ''): v for k, v in d.items()} 
                              for d in data[param_cols].to_dict(orient='records')]
          
    else:  
        data['parameter'] = [{} for _ in range(len(data))]

    data = data[keep_cols + ['parameter']]
    datas.append(data)

df = pd.concat(datas, ignore_index=True)

# sort dataframe
df = df.sort_values(
    by=['symptom', 'model', 'feature'], 
    key=lambda x: x.map({v: i for i, v in enumerate(symptoms if x.name=='symptom' else models if x.name=='model' else features)})
).reset_index(drop=True)

# write into an excel file
with pd.ExcelWriter('results/summary.xlsx', engine='openpyxl') as writer:
    for symptom in symptoms:
        sub_df = df[df['symptom'] == symptom]
        sub_df.to_excel(writer, sheet_name=symptom, index=False) 

# format
wb = load_workbook('results/summary.xlsx')
for sheet in tqdm(symptoms):
    ws = wb[sheet]

    float_cols = []
    for idx, col_name in enumerate(df.columns):
        col_letter = ws.cell(row=1, column=idx+1).column_letter
        if 'rmse' in col_name or 'r2' in col_name:
            float_cols.append(col_letter)
    format_sheet_openpyxl(ws, float_cols=float_cols)

wb.save('results/summary.xlsx')



















